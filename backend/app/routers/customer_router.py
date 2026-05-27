from datetime import datetime
from io import BytesIO
import unicodedata

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import SessionLocal
from app.models.customer_model import Customer
from app.models.plan_model import Plan

from app.schemas.customer_schema import (
    CustomerCreate,
    CustomerResponse,
)

from app.auth.dependencies import require_roles

from app.services.mikrotik_service import (
    create_pppoe_secret,
    update_pppoe_secret,
    disable_pppoe_secret,
    enable_pppoe_secret,
    remove_pppoe_active,
    get_pppoe_secrets,
)

router = APIRouter()


class CustomerImportFromMikrotik(BaseModel):
    pppoe_username: str
    remote_address: str | None = None
    mac_cpe: str | None = None
    router_id: int | None = None
    plan_id: int | None = None
    zone: str | None = None


def get_db():
    db = SessionLocal()

    try:
        yield db
    finally:
        db.close()


def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))

    text = (
        text.replace(" ", "_")
        .replace("-", "_")
        .replace(".", "_")
        .replace("/", "_")
    )

    return text


def clean_value(value):
    if value is None:
        return ""

    return str(value).strip()


def get_excel_value(row, header_index, aliases):
    for alias in aliases:
        key = normalize_text(alias)

        if key in header_index:
            value = row[header_index[key]]
            return clean_value(value)

    return ""


def update_if_not_empty(customer_item, field, value):
    if value is not None and str(value).strip() != "":
        setattr(customer_item, field, value)


def get_customer_profile(db: Session, plan_id):
    if not plan_id:
        return "default"

    plan = (
        db.query(Plan)
        .filter(Plan.id == plan_id)
        .first()
    )

    if not plan:
        return "default"

    return plan.profile or plan.name or "default"


def find_plan_id(db: Session, plan_value: str):
    if not plan_value:
        return None

    plan_value = str(plan_value).strip()

    if plan_value.isdigit():
        return int(plan_value)

    matching_plan = (
        db.query(Plan)
        .filter(
            (Plan.name == plan_value)
            | (Plan.profile == plan_value)
        )
        .first()
    )

    if matching_plan:
        return matching_plan.id

    return None


@router.get("/customers", response_model=list[CustomerResponse])
def get_customers(
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
            "operador",
            "cobrador",
        ])
    ),
):
    return db.query(Customer).all()


@router.post("/customers", response_model=CustomerResponse)
def create_customer(
    customer: CustomerCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    new_customer = Customer(**customer.model_dump())

    db.add(new_customer)
    db.commit()
    db.refresh(new_customer)

    if customer.pppoe_username and customer.pppoe_password:
        try:
            profile = get_customer_profile(db, customer.plan_id)

            class Secret:
                name = customer.pppoe_username
                password = customer.pppoe_password
                service = "pppoe"
                profile = profile
                remote_address = customer.remote_address

            create_pppoe_secret(Secret)

            if customer.status == "suspended":
                disable_pppoe_secret(customer.pppoe_username)
                remove_pppoe_active(customer.pppoe_username)

        except Exception as e:
            print("Error sincronizando PPPoE en MikroTik:", e)

    return new_customer


@router.put(
    "/customers/{customer_id}",
    response_model=CustomerResponse,
)
def update_customer(
    customer_id: int,
    customer: CustomerCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    customer_item = (
        db.query(Customer)
        .filter(Customer.id == customer_id)
        .first()
    )

    if not customer_item:
        raise HTTPException(
            status_code=404,
            detail="Cliente no encontrado",
        )

    old_pppoe_username = customer_item.pppoe_username

    data = customer.model_dump()

    for key, value in data.items():
        setattr(customer_item, key, value)

    db.commit()
    db.refresh(customer_item)

    if customer.pppoe_username and customer.pppoe_password:
        try:
            profile = get_customer_profile(db, customer.plan_id)

            update_pppoe_secret(
                old_name=old_pppoe_username or customer.pppoe_username,
                name=customer.pppoe_username,
                password=customer.pppoe_password,
                profile=profile,
                remote_address=customer.remote_address,
                disabled=customer.status == "suspended",
            )

            if customer.status == "suspended":
                remove_pppoe_active(customer.pppoe_username)

        except Exception as e:
            print("Error actualizando PPPoE en MikroTik:", e)

    return customer_item


@router.put(
    "/customers/{customer_id}/suspend",
    response_model=CustomerResponse,
)
def suspend_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    customer_item = (
        db.query(Customer)
        .filter(Customer.id == customer_id)
        .first()
    )

    if not customer_item:
        raise HTTPException(
            status_code=404,
            detail="Cliente no encontrado",
        )

    customer_item.status = "suspended"

    if customer_item.pppoe_username:
        try:
            disable_pppoe_secret(customer_item.pppoe_username)
            remove_pppoe_active(customer_item.pppoe_username)

        except Exception as e:
            print("Error suspendiendo PPPoE:", e)

    db.commit()
    db.refresh(customer_item)

    return customer_item


@router.put(
    "/customers/{customer_id}/activate",
    response_model=CustomerResponse,
)
def activate_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    customer_item = (
        db.query(Customer)
        .filter(Customer.id == customer_id)
        .first()
    )

    if not customer_item:
        raise HTTPException(
            status_code=404,
            detail="Cliente no encontrado",
        )

    customer_item.status = "active"

    if customer_item.pppoe_username:
        try:
            enable_pppoe_secret(customer_item.pppoe_username)

        except Exception as e:
            print("Error activando PPPoE:", e)

    db.commit()
    db.refresh(customer_item)

    return customer_item


@router.post(
    "/customers/import-from-mikrotik",
    response_model=CustomerResponse,
)
def import_customer_from_mikrotik(
    data: CustomerImportFromMikrotik,
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    existing_customer = (
        db.query(Customer)
        .filter(Customer.pppoe_username == data.pppoe_username)
        .first()
    )

    if existing_customer:
        return existing_customer

    new_customer = Customer(
        name=data.pppoe_username,
        last_name="",
        dni="",
        email="",
        external_id="",
        address="",
        locality="",
        city="",
        postal_code="",
        phone="",
        contract_type="internet",

        pppoe_username=data.pppoe_username,
        pppoe_password="",
        remote_address=data.remote_address or "",
        local_address="",
        mac_cpe=data.mac_cpe or "",
        coordinates="",
        router_id=data.router_id,
        zone=data.zone or "",
        plan_id=data.plan_id,

        status="active",
        notes=(
            f"Cliente importado desde MikroTik el "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}. "
            f"Usuario PPPoE: {data.pppoe_username}. "
            f"IP activa: {data.remote_address or '-'}. "
            f"MAC/Caller ID: {data.mac_cpe or '-'}."
        ),

        billing_type="prepaid",
        invoice_day="",
        payment_day="",
        cut_day="",
    )

    db.add(new_customer)
    db.commit()
    db.refresh(new_customer)

    return new_customer


@router.post("/customers/import-all-from-mikrotik")
def import_all_customers_from_mikrotik(
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    try:
        secrets = get_pppoe_secrets()
    except Exception as e:
        return {
            "status": "error",
            "message": "No se pudo leer PPPoE Secrets desde MikroTik",
            "error": str(e),
        }

    imported = 0
    skipped = 0
    errors = []

    for secret in secrets:
        try:
            username = str(secret.get("name", "")).strip()

            if not username:
                skipped += 1
                continue

            existing_customer = (
                db.query(Customer)
                .filter(Customer.pppoe_username == username)
                .first()
            )

            if existing_customer:
                skipped += 1
                continue

            password = str(secret.get("password", "") or "")
            profile = str(secret.get("profile", "") or "")
            remote_address = str(secret.get("remote-address", "") or "")
            disabled = str(secret.get("disabled", "false")).lower()

            status = (
                "suspended"
                if disabled in ["true", "yes"]
                else "active"
            )

            matching_plan = None

            if profile:
                matching_plan = (
                    db.query(Plan)
                    .filter(
                        (Plan.profile == profile)
                        | (Plan.name == profile)
                    )
                    .first()
                )

            new_customer = Customer(
                name=username,
                last_name="",
                dni="",
                email="",
                external_id="",
                address="",
                locality="",
                city="",
                postal_code="",
                phone="",
                contract_type="internet",

                pppoe_username=username,
                pppoe_password=password,
                remote_address=remote_address,
                local_address="",
                mac_cpe="",
                coordinates="",
                router_id=None,
                zone="",
                plan_id=matching_plan.id if matching_plan else None,

                status=status,
                notes=(
                    "Cliente importado automáticamente desde MikroTik PPPoE Secrets. "
                    f"Profile MikroTik: {profile or '-'}. "
                    f"Remote Address: {remote_address or '-'}. "
                    f"Estado MikroTik: "
                    f"{'disabled' if status == 'suspended' else 'enabled'}."
                ),

                billing_type="prepaid",
                invoice_day="",
                payment_day="",
                cut_day="",
            )

            db.add(new_customer)
            imported += 1

        except Exception as e:
            errors.append({
                "secret": secret.get("name", "sin_nombre"),
                "error": str(e),
            })

    db.commit()

    return {
        "status": "ok",
        "message": "Importación desde MikroTik finalizada",
        "total_mikrotik_secrets": len(secrets),
        "imported": imported,
        "skipped_existing": skipped,
        "errors": errors,
    }


@router.post("/customers/import-excel")
async def import_customers_from_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "tecnico",
        ])
    ),
):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser .xlsx o .xlsm",
        )

    content = await file.read()

    try:
        workbook = load_workbook(
            filename=BytesIO(content),
            data_only=True,
        )
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"No se pudo leer el Excel: {str(e)}",
        )

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    if not rows or len(rows) < 2:
        raise HTTPException(
            status_code=400,
            detail="El Excel no tiene datos suficientes",
        )

    headers = rows[0]
    header_index = {}

    for index, header in enumerate(headers):
        if header:
            header_index[normalize_text(header)] = index

    imported = 0
    updated_existing = 0
    skipped_empty = 0
    errors = []

    for row_number, row in enumerate(rows[1:], start=2):
        try:
            external_id = get_excel_value(
                row,
                header_index,
                [
                    "ID",
                    "id",
                    "external_id",
                    "id_externo",
                    "codigo",
                    "codigo_cliente",
                ],
            )

            remote_address = get_excel_value(
                row,
                header_index,
                [
                    "Ip",
                    "IP",
                    "ip",
                    "remote_address",
                    "remote_ip",
                    "ip_cliente",
                    "ip_pppoe",
                ],
            )

            pppoe_password = get_excel_value(
                row,
                header_index,
                [
                    "Password PPPoE",
                    "password_pppoe",
                    "pppoe_password",
                    "password",
                    "clave",
                    "contraseña",
                    "contrasena",
                ],
            )

            name = get_excel_value(
                row,
                header_index,
                [
                    "Nombre",
                    "nombre",
                    "cliente",
                    "razon_social",
                    "razon",
                ],
            )

            pppoe_username = get_excel_value(
                row,
                header_index,
                [
                    "Usuario",
                    "usuario",
                    "pppoe_username",
                    "usuario_pppoe",
                    "user",
                    "secret",
                    "nombre_secret",
                    "nombre_usuario",
                    "cliente_pppoe",
                ],
            )

            status_raw = get_excel_value(
                row,
                header_index,
                [
                    "Estado",
                    "estado",
                    "status",
                ],
            ).lower()

            plan_value = get_excel_value(
                row,
                header_index,
                [
                    "Plan Internet",
                    "plan_internet",
                    "plan",
                    "plan_id",
                    "id_plan",
                    "profile",
                    "perfil",
                ],
            )

            address = get_excel_value(
                row,
                header_index,
                [
                    "Dirección",
                    "Direccion",
                    "direccion",
                    "address",
                    "domicilio",
                ],
            )

            locality = get_excel_value(
                row,
                header_index,
                [
                    "Barrio/Localidad",
                    "barrio_localidad",
                    "barrio",
                    "localidad",
                    "locality",
                ],
            )

            phone = get_excel_value(
                row,
                header_index,
                [
                    "Telefono",
                    "Teléfono",
                    "telefono",
                    "phone",
                    "celular",
                    "whatsapp",
                ],
            )

            installation_date = get_excel_value(
                row,
                header_index,
                [
                    "Fecha Instalación",
                    "Fecha Instalacion",
                    "fecha_instalacion",
                    "fecha instalación",
                    "fecha instalacion",
                    "instalacion",
                    "installation_date",
                ],
            )

            comments = get_excel_value(
                row,
                header_index,
                [
                    "Comentarios",
                    "comentarios",
                    "notes",
                    "notas",
                    "observaciones",
                ],
            )

            plan_price = get_excel_value(
                row,
                header_index,
                [
                    "Plan Precio",
                    "plan_precio",
                    "precio_plan",
                    "precio",
                ],
            )

            additional_info = get_excel_value(
                row,
                header_index,
                [
                    "Información Adicional",
                    "Informacion Adicional",
                    "informacion_adicional",
                    "info_adicional",
                    "adicional",
                ],
            )

            if not pppoe_username and not external_id:
                skipped_empty += 1
                continue

            if status_raw in [
                "suspendido",
                "suspended",
                "disabled",
                "deshabilitado",
                "cortado",
            ]:
                status = "suspended"
            elif status_raw in [
                "inactivo",
                "inactive",
            ]:
                status = "inactive"
            elif status_raw in [
                "activo",
                "active",
                "enabled",
                "habilitado",
            ]:
                status = "active"
            else:
                status = ""

            plan_id = find_plan_id(db, plan_value)

            notes_parts = []

            if comments:
                notes_parts.append(f"Comentarios: {comments}")

            if installation_date:
                notes_parts.append(f"Fecha instalación: {installation_date}")

            if plan_price:
                notes_parts.append(f"Plan precio: {plan_price}")

            if additional_info:
                notes_parts.append(f"Información adicional: {additional_info}")

            notes_from_excel = "\n".join(notes_parts).strip()

            existing_customer = None

            if pppoe_username:
                existing_customer = (
                    db.query(Customer)
                    .filter(Customer.pppoe_username == pppoe_username)
                    .first()
                )

            if not existing_customer and external_id:
                existing_customer = (
                    db.query(Customer)
                    .filter(Customer.external_id == external_id)
                    .first()
                )

            if existing_customer:
                update_if_not_empty(existing_customer, "external_id", external_id)
                update_if_not_empty(existing_customer, "remote_address", remote_address)
                update_if_not_empty(existing_customer, "pppoe_password", pppoe_password)
                update_if_not_empty(existing_customer, "name", name)
                update_if_not_empty(existing_customer, "pppoe_username", pppoe_username)
                update_if_not_empty(existing_customer, "status", status)
                update_if_not_empty(existing_customer, "address", address)
                update_if_not_empty(existing_customer, "locality", locality)
                update_if_not_empty(existing_customer, "zone", locality)
                update_if_not_empty(existing_customer, "phone", phone)

                if plan_id:
                    existing_customer.plan_id = plan_id

                if notes_from_excel:
                    existing_notes = existing_customer.notes or ""

                    import_note = (
                        f"\n\nActualizado desde Excel el "
                        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}.\n"
                        f"{notes_from_excel}"
                    )

                    existing_customer.notes = (existing_notes + import_note).strip()

                updated_existing += 1

            else:
                new_customer = Customer(
                    name=name or pppoe_username or external_id,
                    last_name="",
                    dni="",
                    email="",
                    external_id=external_id,
                    address=address,
                    locality=locality,
                    city="",
                    postal_code="",
                    phone=phone,
                    contract_type="internet",

                    pppoe_username=pppoe_username,
                    pppoe_password=pppoe_password,
                    remote_address=remote_address,
                    local_address="",
                    mac_cpe="",
                    coordinates="",
                    router_id=None,
                    zone=locality,
                    plan_id=plan_id,

                    status=status or "active",
                    notes=(
                        notes_from_excel
                        or f"Cliente importado desde Excel el {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}."
                    ),

                    billing_type="prepaid",
                    invoice_day="",
                    payment_day="",
                    cut_day="",
                )

                db.add(new_customer)
                imported += 1

        except Exception as e:
            errors.append({
                "row": row_number,
                "error": str(e),
            })

    db.commit()

    return {
        "status": "ok",
        "message": "Importación Excel finalizada",
        "filename": file.filename,
        "imported": imported,
        "updated_existing": updated_existing,
        "skipped_existing": 0,
        "skipped_empty": skipped_empty,
        "errors": errors,
    }


@router.delete("/customers/{customer_id}")
def delete_customer(
    customer_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
        ])
    ),
):
    customer_item = (
        db.query(Customer)
        .filter(Customer.id == customer_id)
        .first()
    )

    if not customer_item:
        raise HTTPException(
            status_code=404,
            detail="Cliente no encontrado",
        )

    if customer_item.pppoe_username:
        try:
            disable_pppoe_secret(customer_item.pppoe_username)
            remove_pppoe_active(customer_item.pppoe_username)

        except Exception as e:
            print("Error eliminando conexión activa PPPoE:", e)

    db.delete(customer_item)
    db.commit()

    return {
        "message": "Cliente eliminado"
    }