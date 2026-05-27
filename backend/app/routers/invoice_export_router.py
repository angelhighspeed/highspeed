from io import BytesIO
from datetime import date, datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import SessionLocal
from app.auth.dependencies import require_roles
from app.models.invoice_model import Invoice
from app.models.customer_model import Customer


router = APIRouter()


def get_db():
    db = SessionLocal()

    try:
        yield db
    finally:
        db.close()


def format_date(value):
    if value is None:
        return ""

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    return str(value)


def format_status(status):
    if status == "paid":
        return "Pagada"

    if status == "pending":
        return "Pendiente"

    return status or ""


@router.get("/invoices/export-excel")
def export_invoices_excel(
    status: str | None = Query(None),
    db: Session = Depends(get_db),
    current_user: dict = Depends(
        require_roles([
            "admin",
            "cobrador",
        ])
    ),
):
    query = db.query(Invoice)

    if status:
        query = query.filter(Invoice.status == status)

    invoices = query.order_by(Invoice.id.desc()).all()

    customer_ids = list({
        invoice.customer_id
        for invoice in invoices
        if invoice.customer_id
    })

    customers = (
        db.query(Customer)
        .filter(Customer.id.in_(customer_ids))
        .all()
        if customer_ids
        else []
    )

    customers_by_id = {
        customer.id: customer
        for customer in customers
    }

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Facturas"

    headers = [
        "Factura ID",
        "Cliente ID",
        "Cliente",
        "Usuario PPPoE",
        "IP",
        "Teléfono",
        "Zona",
        "Monto",
        "Vencimiento",
        "Estado",
    ]

    sheet.append(headers)

    header_fill = PatternFill(
        start_color="2563EB",
        end_color="2563EB",
        fill_type="solid",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for invoice in invoices:
        customer = customers_by_id.get(invoice.customer_id)

        customer_name = ""

        if customer:
            customer_name = f"{customer.name or ''} {customer.last_name or ''}".strip()

        sheet.append([
            invoice.id,
            invoice.customer_id,
            customer_name or "-",
            customer.pppoe_username if customer else "-",
            customer.remote_address if customer else "-",
            customer.phone if customer else "-",
            customer.zone if customer else "-",
            float(invoice.amount or 0),
            format_date(invoice.due_date),
            format_status(invoice.status),
        ])

    widths = {
        1: 12,
        2: 12,
        3: 32,
        4: 22,
        5: 18,
        6: 18,
        7: 18,
        8: 14,
        9: 16,
        10: 16,
    }

    for index, width in widths.items():
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    sheet.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = "facturas"

    if status == "pending":
        filename = "facturas_pendientes"
    elif status == "paid":
        filename = "facturas_pagadas"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}.xlsx"'
        },
    )